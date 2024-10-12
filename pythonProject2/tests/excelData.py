import copy
import csv

import openpyxl
li=list()
Dict=dict()
book = openpyxl.load_workbook("C:\\Users\\siddh\\Downloads\\Data.xlsx")
sheet = book.active
cell=sheet.cell(row=1, column=2)
print (cell.value)

sheet.cell(row=2, column=2).value = "Rahul"
#book.save("C:\\Users\\siddh\\Downloads\\Data.xlsx")
print(sheet.cell(row=2, column=2).value)

print(sheet.max_row)


for i in range(2, sheet.max_row+1):
    #if sheet.cell(row=i, column=1).value == "Testcase2":
    for j in range(2, sheet.max_column+1):
            #print(sheet.cell(row=i, column=j).value)
        Dict[sheet.cell(row=1, column=j).value]=sheet.cell(row=i, column=j).value
        dict1= copy.deepcopy(Dict)
    li.append(dict1)
print(li)

if __name__ == '__main__':
    pass
